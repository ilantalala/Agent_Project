from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path

# Excel file will be created next to the backend code
FILE_PATH = Path(__file__).resolve().parent / "conversations.xlsx"

HEADERS = ["Timestamp", "Name", "Phone", "Conversation"]

def log_conversation(name: str, phone: str, conversation: str) -> None:
    if FILE_PATH.exists():
        wb = load_workbook(FILE_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Conversations"
        ws.append(HEADERS)

    ws.append([
        datetime.utcnow().isoformat(),
        name,
        phone,
        conversation
    ])

    wb.save(FILE_PATH)
